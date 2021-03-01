#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Feb 12 08:30:26 2021

@author: oguzhanozdmr
"""

import argparse
import os.path
from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill
from openpyxl import load_workbook


new_workbook = Workbook()
new_workbook.remove(new_workbook['Sheet'])
case_sensitive = False
just_first_color = '5CFF5C'
just_second_color = 'FFC55C'
diff_color = 'FFFF00'


def is_file(paths: str):
    isFile = os.path.isfile(paths[0]) and os.path.isfile(paths[1])
    return isFile


def read_excel(path: str):
    wb = load_workbook(filename=path, read_only=True)
    return wb


def read_sheet(ex_workbook):
    excel_sheet_names = ex_workbook.sheetnames
    sheets = {}
    for sheet in excel_sheet_names:
        sheets[sheet] = ex_workbook[sheet]
    return sheets


def is_equal(first_value: str, second_value: str) -> bool:
    isEqual = False
    if case_sensitive:
        if first_value == second_value:
            isEqual = True
    else:
        if first_value.lower() == second_value.lower():
            isEqual = True
    return isEqual


def compare_to_value(first_value: str, second_value: str,
                     current_sheet, row: int, column: int) -> None:
    if first_value and second_value:
        if is_equal(first_value, second_value):
            current_sheet.cell(row, column).value = first_value
        else:
            value = f'{first_value} || {second_value}'
            current_sheet.cell(row, column).value = value
            style = PatternFill(fgColor=diff_color, fill_type='solid')
            current_sheet.cell(row, column).fill = style
    elif first_value:
        current_sheet.cell(row, column).value = first_value
        style = PatternFill(fgColor=just_first_color, fill_type='solid')
        current_sheet.cell(row, column).fill = style
    elif second_value:
        current_sheet.cell(row, column).value = second_value
        style = PatternFill(fgColor=just_second_color, fill_type='solid')
        current_sheet.cell(row, column).fill = style


def compare_excel(first_file, second_file,
                  filepath: str = 'match.xlsx'):

    first_sheets = read_sheet(first_file)
    second_sheets = read_sheet(second_file)

    for sheet in first_sheets:

        first_max_row = first_sheets[sheet].max_row
        first_max_column = first_sheets[sheet].max_column

        new_workbook.create_sheet(sheet)
        current_sheet = new_workbook[sheet]

        if sheet not in second_sheets:
            for row in range(1, first_max_row + 1):
                for column in range(1, first_max_column + 1):
                    value = first_sheets[sheet].cell(row, column).value
                    current_sheet.cell(row, column).value = value
                    style = PatternFill(fgColor=just_first_color,
                                        fill_type='solid')
                    current_sheet.cell(row, column).fill = style

        else:
            second_max_row = second_sheets[sheet].max_row
            second_max_column = second_sheets[sheet].max_column

            max_row = max(first_max_row, second_max_row)
            max_column = max(first_max_column, second_max_column)

            for row in range(1, max_row + 1):
                for column in range(1, max_column + 1):
                    first_value = first_sheets[sheet].cell(row, column).value
                    second_value = second_sheets[sheet].cell(row, column).value
                    compare_to_value(first_value, second_value, current_sheet,
                                     row, column)

    for sheet in second_sheets:
        if sheet not in first_sheets:
            new_workbook.create_sheet(sheet)
            max_row = second_sheets[sheet].max_row
            max_column = second_sheets[sheet].max_column

            for row in range(1, max_row + 1):
                for column in range(1, max_column + 1):
                    value = second_sheets[sheet].cell(row, column).value
                    new_workbook[sheet].cell(row, column).value = value
                    style = PatternFill(fgColor=just_second_color,
                                        fill_type='solid')
                    current_sheet.cell(row, column).fill = style

    new_workbook.save(filepath)


def start_command_line():
    help1 = "Xlsx file is compotared to excel file"
    my_parser = argparse.ArgumentParser(description=help1)
    my_parser.version = 1

    help1 = "Xlsx files to be compared"
    my_parser.add_argument('files', nargs='+', help=help1)

    help1 = "The case sensitive is defualt False"
    my_parser.add_argument('-cs', '-casesensitive',
                           action='store_true',
                           help=help1)
    data = my_parser.parse_args()
    return data


def main():
    parser = start_command_line()
    case_sensitive = parser.cs
    if len(parser.files) < 2:
        print('You must enter two files path')
    elif is_file(parser.files):
        first_path = parser.files[0]
        second_path = parser.files[1]
        first_excel = read_excel(first_path)
        second_excel = read_excel(second_path)
        compare_excel(first_excel, second_excel)
        first_excel.close()
        second_excel.close()
        print('Done!')
    else:
        print('Not found files')


if __name__ == '__main__':
    main()
